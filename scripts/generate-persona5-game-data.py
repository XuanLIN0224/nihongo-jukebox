#!/usr/bin/env python3
"""Generate Persona 5 game-study data from the user-provided workbook.

The workbook is used as a data source only. The generated JSON keeps every
message block found in the Original sheet and adds Anki furigana when available.
Chinese is an automatic local draft so every line has jp/en/zh fields offline.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


FILE_RE = re.compile(r"^[A-Za-z0-9_./-]+\.bf$")
ID_RE = re.compile(r"^Id:\s*(\d+)$", re.I)
DUMMY_RE = re.compile(r"^(?:Msg_)?DUMMY|^Dummy\s+", re.I)
JP_RE = re.compile(r"[ぁ-んァ-ヶー一-龥々]")


@dataclass
class Block:
  file: str
  msg_id: str
  key: str
  lines: list[str] = field(default_factory=list)


def clean_cell(value: object) -> str:
  if value is None:
    return ""
  return str(value).replace("\u3000", " ").strip()


def block_key(file: str, msg_id: str, key: str) -> str:
  return f"{file}_Id:_{msg_id}_{key}"


def parse_side(rows: list[tuple[object, ...]], column_index: int) -> dict[str, Block]:
  blocks: dict[str, Block] = {}
  current_file = ""
  current_id = ""
  current_key = ""
  lines: list[str] = []

  def commit() -> None:
    nonlocal lines
    if current_file and current_id and current_key:
      key = block_key(current_file, current_id, current_key)
      if key not in blocks:
        blocks[key] = Block(file=current_file, msg_id=current_id, key=current_key, lines=[])
      blocks[key].lines.extend(line for line in lines if line)
    lines = []

  for row in rows:
    value = clean_cell(row[column_index] if column_index < len(row) else "")
    if not value:
      continue
    if FILE_RE.match(value):
      commit()
      current_file = value
      current_id = ""
      current_key = ""
      lines = []
      continue
    id_match = ID_RE.match(value)
    if id_match:
      commit()
      current_id = id_match.group(1)
      current_key = ""
      lines = []
      continue
    if current_file and current_id and not current_key:
      current_key = value
      lines = []
      continue
    if current_file and current_id and current_key:
      lines.append(value)

  commit()
  return blocks


def load_anki(path: Path) -> dict[str, dict[str, str]]:
  wb = load_workbook(path, read_only=True, data_only=True)
  if "Anki" not in wb.sheetnames:
    return {}
  ws = wb["Anki"]
  rows: dict[str, dict[str, str]] = {}
  for raw in ws.iter_rows(values_only=True):
    values = [clean_cell(value) for value in raw[:4]]
    if not values or not values[0]:
      continue
    key = values[0]
    jp = values[1]
    en = values[2]
    furigana = values[3]
    if not jp and JP_RE.search(en):
      jp, en = en, ""
    rows[key] = {"japanese": jp, "english": en, "furigana": furigana}
  return rows


def split_speaker(lines: list[str]) -> tuple[str, str]:
  if lines and re.search(r"[:：]$", lines[0]) and len(lines[0]) <= 48:
    return lines[0].rstrip(":："), "\n".join(lines[1:]).strip()
  return "", "\n".join(lines).strip()


def compact_text(value: str) -> str:
  return re.sub(r"[ \t]+", " ", value).strip()


EN_PHRASES = [
  ("What exactly is it?How did that barrier open up?Why are you so interested in it?I know all I need to know.Nah, I'm good.", "那到底是什么？那道屏障是怎么打开的？你为什么这么在意它？该知道的我都知道了。不用了，我没问题。"),
  ("Well, it's the collective Palace of the general public.", "嗯，那是大众的集体殿堂。"),
  ("Deep inside each and every human's mind is an obscured thought process known as the unconscious.", "每个人内心深处都有一种被称为无意识的隐蔽思考过程。"),
  ("These thought processes then meld together and form what I like to call the collective unconscious.", "这些思考过程会融合在一起，形成我称作集体无意识的东西。"),
  ("Mementos is a shared cognitive world created by the joint distortions of society as a whole.", "印象空间是由整个社会共同扭曲所创造出的共享认知世界。"),
  ("You can just think of it as the big Palace made from everyone's hearts though.", "你也可以简单把它理解成由大家的心构成的巨大殿堂。"),
  ("I would bet it was because of the shift in public opinion caused by Kamoshida's change of heart.", "我猜是鸭志田改心引发的舆论变化造成的。"),
  ("Those kinds of big changes in the tide will surely have an impact on what goes on in Mementos.", "那种巨大的潮流变化肯定会影响印象空间内部的情况。"),
  ("We'll need to take down another target in order for that to happen though", "不过要做到这一点，我们还得解决下一个目标"),
  ("I have a feeling that something deep within that place has caused me to take my current form.", "我有种感觉，是那个地方深处的某种东西让我变成了现在的样子。"),
  ("That's why I'm so invested in reaching the depths of Mementos.", "所以我才这么想抵达印象空间的最深处。"),
  ("If we can figure out the cause of my distorted form, I should be able to go back to normal", "如果能弄清我这副扭曲形态的原因，我应该就能恢复原状"),
  ("Either way, I'll be counting on you guys to help me out.", "不管怎样，我就指望你们帮我了。"),
  ("Oh, and it's important to note that both Mementos and the Palaces are affected by real world happenings.", "哦，还有一点很重要：印象空间和殿堂都会受到现实世界事件的影响。"),
  ("Weather's a great example of that.", "天气就是很好的例子。"),
  ("I mean, weather affects your mood, doesn't it?", "我是说，天气会影响你的心情，不是吗？"),
  ("Same goes for Mementos.", "印象空间也是一样。"),
  ("You should try heading in on a day with bad weather.", "你可以试着在天气不好的日子进去看看。"),
  ("I'll explain in more detail then.", "到时候我再详细说明。"),
  ("Mishima's going to be a great source of info for us.", "三岛会成为我们很好的情报来源。"),
  ("Make sure you form a strong bond with him, OK?", "一定要和他建立稳固的关系，明白吗？"),
  ("Next Wednesday is the start of exams", "下周三就要开始考试了"),
  ("Shouldn't you be studying?", "你不用学习吗？"),
  ("I know exams are important, but don't forget about Mementos, OK?", "我知道考试很重要，但也别忘了印象空间，明白吗？"),
  ("Your exams are finally starting tomorrow.", "你的考试明天终于开始了。"),
  ("I can't wait to see how you end up doing.", "我很期待看看你最后会考得怎么样。"),
  ("I'll be sure to escort Lady Ann tomorrow!", "明天我一定会好好护送杏大人！"),
  ("How dare you deny King Kamoshida's love, you selfish lass", "你竟敢否认鸭志田王的爱，你这个自私的小姑娘"),
  ("King Kamoshida's", "鸭志田王的"),
  ("He's trying to do something", "他想做什么"),
  ("How dare you", "你竟敢"),
  ("King Kamoshida", "鸭志田王"),
  ("Pay for this insolence with your life", "用你的命为这份无礼付出代价"),
  ("that guy's nothing but a pathetic loser", "那家伙不过是个可悲的废物"),
  ("there would be a woman", "竟然会有女人"),
  ("could stand up to", "敢反抗"),
  ("outside of school", "在学校外面"),
  ("collective unconscious", "集体无意识"),
  ("Collective unconscious", "集体无意识"),
  ("general public", "大众"),
  ("public opinion", "公众舆论"),
  ("change of heart", "改心"),
  ("big changes", "巨大的变化"),
  ("in other words", "换句话说"),
  ("In other words", "换句话说"),
  ("take down another target", "解决下一个目标"),
  ("current form", "现在的样子"),
  ("go back to normal", "恢复原状"),
  ("counting on you", "指望你"),
  ("bad weather", "坏天气"),
  ("great source of info", "很好的情报来源"),
  ("strong bond", "稳固的关系"),
  ("start of exams", "考试开始"),
  ("finally starting tomorrow", "明天终于开始"),
  ("Lady Ann", "杏大人"),
  ("Make sure", "一定要"),
  ("be slaughtered", "被杀掉"),
  ("at this rate", "照这样下去"),
  ("You'll be slaughtered at this rate", "照这样下去你会被杀掉"),
  ("Phantom Thieves", "怪盗团"),
  ("Treasure", "秘宝"),
  ("Palace", "殿堂"),
  ("Shadow", "暗影"),
  ("Shadows", "暗影"),
  ("Persona", "人格面具"),
  ("Metaverse", "异世界"),
  ("Cognitive", "认知"),
  ("cognitive", "认知"),
  ("calling card", "预告信"),
  ("Calling card", "预告信"),
  ("Velvet Room", "天鹅绒房间"),
  ("Confidant", "协助者"),
  ("Joker", "Joker"),
  ("Morgana", "摩尔加纳"),
  ("Ryuji", "龙司"),
  ("Ann", "杏"),
  ("Yusuke", "祐介"),
  ("Makoto", "真"),
  ("Futaba", "双叶"),
  ("Haru", "春"),
  ("Akechi", "明智"),
  ("Sae", "冴"),
  ("Sojiro", "惣治郎"),
  ("Igor", "伊戈尔"),
  ("Arsene", "亚森"),
  ("you are", "你是"),
  ("you're", "你是"),
  ("You are", "你是"),
  ("You're", "你是"),
  ("I am", "我是"),
  ("I'm", "我是"),
  ("we are", "我们是"),
  ("We're", "我们是"),
  ("don't", "不要"),
  ("Don't", "不要"),
  ("can't", "不能"),
  ("Can't", "不能"),
  ("won't", "不会"),
  ("Won't", "不会"),
  ("have to", "必须"),
  ("need to", "需要"),
  ("Let's", "我们"),
  ("let's", "我们"),
  ("I have a feeling", "我有种感觉"),
  ("That's why", "所以"),
  ("Either way", "不管怎样"),
  ("I'll be", "我会"),
  ("I would bet", "我猜"),
  ("we might be able to", "我们也许能够"),
  ("be able to", "能够"),
  ("going to be", "会成为"),
  ("I know", "我知道"),
  ("I can't wait", "我很期待"),
  ("how you end up doing", "你最后会表现得怎样"),
  ("I mean", "我是说"),
  ("doesn't it", "不是吗"),
  ("Same goes for", "同样也适用于"),
  ("Thank you", "谢谢"),
  ("thanks", "谢谢"),
  ("sorry", "对不起"),
  ("Sorry", "对不起"),
  ("Please", "请"),
  ("please", "请"),
]

EN_WORDS = {
  "the": "",
  "a": "",
  "an": "",
  "to": "",
  "of": "的",
  "and": "和",
  "or": "或者",
  "but": "但是",
  "if": "如果",
  "because": "因为",
  "with": "和",
  "without": "没有",
  "this": "这个",
  "that": "那个",
  "these": "这些",
  "those": "那些",
  "here": "这里",
  "there": "那里",
  "now": "现在",
  "then": "然后",
  "why": "为什么",
  "what": "什么",
  "who": "谁",
  "where": "哪里",
  "when": "什么时候",
  "how": "怎么",
  "me": "我",
  "my": "我的",
  "i": "我",
  "you": "你",
  "your": "你的",
  "he": "他",
  "she": "她",
  "we": "我们",
  "they": "他们",
  "it": "它",
  "all": "全部",
  "go": "走",
  "come": "来",
  "get": "得到",
  "take": "拿",
  "give": "给",
  "make": "做",
  "see": "看见",
  "look": "看",
  "hear": "听见",
  "say": "说",
  "tell": "告诉",
  "know": "知道",
  "think": "想",
  "feel": "感觉",
  "want": "想要",
  "can": "可以",
  "will": "会",
  "would": "会",
  "could": "可以",
  "should": "应该",
  "must": "必须",
  "fight": "战斗",
  "kill": "杀",
  "die": "死",
  "enemy": "敌人",
  "enemies": "敌人",
  "friend": "朋友",
  "friends": "朋友",
  "power": "力量",
  "heart": "心",
  "desire": "欲望",
  "desires": "欲望",
  "world": "世界",
  "truth": "真相",
  "justice": "正义",
  "school": "学校",
  "teacher": "老师",
  "student": "学生",
  "police": "警察",
  "case": "案件",
  "room": "房间",
  "money": "钱",
  "time": "时间",
  "day": "天",
  "night": "夜晚",
  "good": "好",
  "bad": "坏",
  "right": "正确",
  "wrong": "错误",
  "real": "真实",
  "ideal": "理想",
  "secret": "秘密",
  "change": "改变",
  "changes": "变化",
  "escape": "逃跑",
  "run": "跑",
  "stop": "停下",
  "wait": "等",
  "help": "帮助",
  "save": "拯救",
  "remember": "记住",
  "forget": "忘记",
  "believe": "相信",
  "understand": "明白",
  "leave": "离开",
  "alive": "活着",
  "free": "自由",
  "choice": "选择",
  "answer": "答案",
  "question": "问题",
  "exactly": "到底",
  "did": "做了",
  "barrier": "屏障",
  "open": "打开",
  "up": "起来",
  "interested": "感兴趣",
  "nah": "不用",
  "well": "嗯",
  "collective": "集体",
  "public": "大众",
  "deep": "深处",
  "inside": "里面",
  "each": "每个",
  "every": "每个",
  "human": "人",
  "humans": "人类",
  "human's": "人的",
  "mind": "心",
  "obscured": "隐蔽的",
  "thought": "想法",
  "process": "过程",
  "processes": "过程",
  "known": "被称为",
  "unconscious": "无意识",
  "meld": "融合",
  "together": "一起",
  "form": "形成",
  "forms": "形态",
  "created": "创造的",
  "joint": "共同的",
  "distortions": "扭曲",
  "society": "社会",
  "whole": "整体",
  "just": "只是",
  "big": "大的",
  "made": "做成的",
  "everyone's": "大家的",
  "hearts": "心",
  "bet": "猜",
  "was": "是",
  "shift": "转变",
  "caused": "造成的",
  "opinion": "舆论",
  "kinds": "种",
  "tide": "潮流",
  "surely": "肯定",
  "impact": "影响",
  "goes": "发生",
  "on": "着",
  "might": "也许",
  "able": "能够",
  "further": "更深处",
  "rouse": "唤起",
  "more": "更多",
  "acknowledgment": "认同",
  "need": "需要",
  "down": "打倒",
  "another": "另一个",
  "target": "目标",
  "happen": "发生",
  "though": "不过",
  "feeling": "感觉",
  "something": "某种东西",
  "within": "内部",
  "place": "地方",
  "has": "已经",
  "invested": "投入",
  "reaching": "抵达",
  "depths": "深处",
  "figure": "弄清楚",
  "out": "出来",
  "cause": "原因",
  "distorted": "扭曲的",
  "normal": "正常",
  "either": "不管",
  "way": "方式",
  "counting": "指望",
  "guys": "你们",
  "both": "两者",
  "affected": "受到影响",
  "happenings": "事件",
  "weather": "天气",
  "great": "很好的",
  "example": "例子",
  "affects": "影响",
  "mood": "心情",
  "same": "同样",
  "try": "试着",
  "heading": "前往",
  "explain": "说明",
  "detail": "细节",
  "next": "下一个",
  "wednesday": "星期三",
  "exams": "考试",
  "studying": "学习",
  "important": "重要",
  "finally": "终于",
  "starting": "开始",
  "tomorrow": "明天",
  "escort": "护送",
  "dare": "敢",
  "deny": "否认",
  "king": "王",
  "love": "爱",
  "selfish": "自私的",
  "lass": "女孩",
  "pay": "付出代价",
  "insolence": "无礼",
  "life": "生命",
  "woman": "女人",
  "stand": "反抗",
  "outside": "外面",
  "guy": "家伙",
  "guy's": "家伙的",
  "nothing": "没有什么",
  "pathetic": "可悲的",
  "loser": "废物",
  "trying": "试图",
  "slaughtered": "屠杀",
  "rate": "情况",
}

JP_EXACT_TRANSLATIONS = [
  (r"何やってんだ.*早く逃げるんだよ", "你在干什么！快逃啊！！"),
  (r"ヤメロ", "住手！！！"),
  (r"覚悟しろ.*理不尽.*ブッ潰してやる", "做好觉悟吧，你们这些家伙……这种不讲理的事，我会亲手粉碎！"),
  (r"邪魔すんな.*バケモノ", "别碍事！怪物！！"),
  (r"追い詰めたぞ.*賊め", "把你逼到绝路了，贼人！"),
]

JP_HINTS = [
  (r"ありがとう", "谢谢。"),
  (r"すみません|ごめん", "对不起。"),
  (r"行く|行け", "走吧。"),
  (r"逃げ", "快逃。"),
  (r"戦", "战斗。"),
  (r"敵", "敌人。"),
  (r"死", "死亡。"),
  (r"力", "力量。"),
  (r"心", "心。"),
  (r"欲望", "欲望。"),
]


def auto_translate_english(text: str, japanese: str) -> str:
  source = compact_text(text.replace("\n", " ")).replace("’", "'")
  if not source:
    for pattern, zh in JP_EXACT_TRANSLATIONS:
      if re.search(pattern, japanese):
        return zh
    for pattern, zh in JP_HINTS:
      if re.search(pattern, japanese):
        return zh
    return "（无英文原文；请参考日文。）"
  if JP_RE.search(source):
    for pattern, zh in JP_EXACT_TRANSLATIONS:
      if re.search(pattern, source):
        return zh
  translated = source
  for en, zh in sorted(EN_PHRASES, key=lambda item: len(item[0]), reverse=True):
    pattern = re.escape(en).replace(r"\ ", r"\s+")
    if re.match(r"^[A-Za-z0-9' -]+$", en):
      pattern = rf"(?<![A-Za-z]){pattern}(?![A-Za-z])"
    translated = re.sub(pattern, zh, translated, flags=re.I)

  def word_repl(match: re.Match[str]) -> str:
    word = match.group(0)
    lower = word.lower()
    return EN_WORDS.get(lower, word)

  translated = re.sub(r"[A-Za-z']+", word_repl, translated)
  translated = re.sub(r"\s+", " ", translated).strip()
  translated = translated.replace(" .", "。").replace(".", "。")
  translated = translated.replace(" ?", "？").replace("?", "？")
  translated = translated.replace(" !", "！").replace("!", "！")
  translated = translated.replace(" ,", "，").replace(",", "，")
  translated = translated.replace(" :", "：")
  return translated or "（自动中文待精修。）"


def stable_id(key: str) -> str:
  safe = re.sub(r"[^A-Za-z0-9_-]+", "-", key).strip("-")
  return f"p5-{safe}"


def main() -> None:
  parser = argparse.ArgumentParser()
  parser.add_argument("workbook", type=Path)
  parser.add_argument("output", type=Path)
  args = parser.parse_args()

  wb = load_workbook(args.workbook, read_only=True, data_only=True)
  ws = wb["Original"]
  rows = list(ws.iter_rows(values_only=True))
  jp_blocks = parse_side(rows, 0)
  en_blocks = parse_side(rows, 4)
  anki = load_anki(args.workbook)

  keys = sorted(set(jp_blocks) | set(en_blocks), key=lambda value: (
    value.split("_Id:_", 1)[0],
    int(re.search(r"_Id:_(\d+)_", value).group(1)) if re.search(r"_Id:_(\d+)_", value) else 0,
    value,
  ))

  lines = []
  file_counts: dict[str, int] = {}
  for index, key in enumerate(keys):
    jp_block = jp_blocks.get(key)
    en_block = en_blocks.get(key)
    template = jp_block or en_block
    if not template:
      continue
    jp_speaker, jp_text = split_speaker(jp_block.lines if jp_block else [])
    en_speaker, en_text = split_speaker(en_block.lines if en_block else [])
    anki_row = anki.get(key, {})
    if anki_row.get("japanese"):
      jp_text = anki_row["japanese"]
    if anki_row.get("english"):
      en_text = anki_row["english"]
    is_dummy = bool(DUMMY_RE.search(template.key) or DUMMY_RE.search(jp_text) or DUMMY_RE.search(en_text))
    line = {
      "id": stable_id(key),
      "sourceKey": key,
      "file": template.file,
      "messageId": template.msg_id,
      "messageKey": template.key,
      "speakerJp": jp_speaker,
      "speakerEn": en_speaker,
      "japanese": compact_text(jp_text),
      "english": compact_text(en_text),
      "zh": auto_translate_english(en_text, jp_text),
      "furigana": compact_text(anki_row.get("furigana", "")),
      "isDummy": is_dummy,
      "order": index,
    }
    lines.append(line)
    file_counts[template.file] = file_counts.get(template.file, 0) + 1

  payload = {
    "id": "persona5",
    "title": "Persona 5",
    "sourceWorkbook": args.workbook.name,
    "generatedAt": "2026-09-01T00:00:00.000Z",
    "lineCount": len(lines),
    "fileCount": len(file_counts),
    "nonDummyLineCount": sum(1 for line in lines if not line["isDummy"]),
    "translationMode": "local-auto-draft",
    "lines": lines,
  }
  args.output.parent.mkdir(parents=True, exist_ok=True)
  args.output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
  print(f"Generated {len(lines)} Persona 5 message cards across {len(file_counts)} files.")
  print(f"Non-dummy cards: {payload['nonDummyLineCount']}.")


if __name__ == "__main__":
  main()
